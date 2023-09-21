from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from os import path
import sys
import sqlite3 as sq
import openpyxl
import random
from PyQt5 import uic
from datetime import datetime as dt
FORM_CLASS,_ = uic.loadUiType(path.join(path.dirname(__file__),'untitled.ui'))




class Mainapp(QMainWindow, FORM_CLASS):
    #QWidget
    total_monay_of_invoice  = 0 
    def __init__(self, parent=None):
        super(Mainapp, self).__init__(parent)
        QMainWindow.__init__(self)
        self.setupUi(self)
        
        
        self.Handel_ui()
        self.Handel_Tabels_Ui()
        self.DB_Management()
        self.Handel_All_btns()
        self.DB_Management()
        self.put_units_in_combobox()
        self.get_data_from_selected_row()
        
        
    #_______________________________________
    def DB_Management(self):
        self.connection  = sq.connect(path.join(path.dirname(__file__), r'DB folder\DB.db' ))   
        self.cursor1 = self.connection.cursor()
             
    def Handel_ui(self):
        self.setWindowTitle( "Inventory Management System" )
        self.setFixedSize( 1358, 870)
        self.tabWidget.setCurrentIndex(0)

    def Handel_Tabels_Ui(self):
        self.Handel_Storge_Table_Ui()
        self.Handel_seelingTab_tabels_Ui()
        self.customers_tables_ui()
        
    def Handel_Storge_Table_Ui(self):
        self.Storge_Table.setColumnWidth(0,270)
        self.Storge_Table.setColumnWidth(1,270)
        self.Storge_Table.setColumnWidth(2,270)
        self.Storge_Table.setColumnWidth(3,270)
        self.Storge_Table.setColumnWidth(4,270)
        
    def Handel_All_btns(self):
        self.Handel_main_buttons()
        self.Handel_Storge_Buttons()
        self.Handel_sell_Buttons()
        self.Handel_invoice_tab_btns()
        self.Handel_customers_tab_btns()
     
     
#*------------------------------------------------------------------------------ Storge  --------------------------------------------
        
    def Handel_Storge_Buttons(self):
        self.storge_tab_add_btn.clicked.connect(self.Storge_Add_func)
        self.storge_tab_search_btn.clicked.connect(self.Storge_Search_func)
        self.storge_tab_update_btn.clicked.connect(self.Storge_Update_func)
        self.storge_tab_showall_btn.clicked.connect(self.Storge_ShowAll_func)
        self.storge_tab_add_ToComboBox_btn_6.clicked.connect(self.add_unit_to_combobox_func) 
     
    def Storge_Add_func(self):
        Product_name , buying_price , sell_price , unit_type, total_num_of_units = self.selling_get_data_from_Ui()
        if Product_name and sell_price and unit_type and total_num_of_units   :
                valid = self.check_if_productname_is_Available( TableName='Storge' , Product_name = Product_name)
                if  valid :
                    self.cursor1.execute(""" insert into Storge ( [Product_name] , [buying_price] , [sell_price] , [unit_type] , [total_num_of_units]) 
                                    values (?,?,?,?,?)""", ( Product_name , buying_price , sell_price , unit_type , total_num_of_units ))
                    self.connection.commit() 
                    QMessageBox.information(self, "i" , "تم أضافه المنتج بنجاح")
                    self.Storge_ShowAll_func()
                else: QMessageBox.warning(self, "i" , "يوجد منتج بنفس الاسم ادخل اسم مختلف")
        else: QMessageBox.warning(self, "i" , "لم تقم بأدخال أحدي البينات التاليه الاسم ,سعر البيع ,الفئه او عدد الوحدات المباعه")
        
    def add_unit_to_combobox_func(self):
        item = str(self.storge_tab_unit_type_comboBox.currentText())
        print(item)
        self.cursor1.execute(""" insert into units ([unit_name] , [c]) values (?,?)""", (item , 1) )
        self.connection.commit()
        self.put_units_in_combobox()
        
    def put_units_in_combobox(self):
        self.cursor1.execute(""" select [unit_name] from units """)
        items  = self.cursor1.fetchall()
        if items:
            for item in items :
                self.storge_tab_unit_type_comboBox.addItem(item[0])
                             
    def Storge_Search_func(self):
        
        name = self.storge_tab_product_name_line.text()
        if name:
            self.cursor1.execute(f"""select * from Storge where Product_name like '%{name}%'""")
            rows = self.cursor1.fetchall()
            self.Storge_Table.clearContents()
            if rows:
                self.Storge_Table.setRowCount(len(rows))
                for i, row in enumerate(rows):
                    for j in range(len(row)):
                        item = QTableWidgetItem(str(row[j]))
                        self.Storge_Table.setItem(i, j, item)
            else:
                QMessageBox.warning(self, "i" , "لا يوجد منتج بهذه المعلومات")
        else:
            QMessageBox.warning(self, "i" , "يجب ادخال اسم المنتج لأتمام عمليه البحث")
    
    def Storge_Update_func(self):
        Product_name , buying_price , sell_price , unit_type, total_num_of_units = self.get_data_from_storge_Ui()
        if Product_name and sell_price and unit_type and total_num_of_units   :
            valid = self.check_if_productname_is_Available( TableName='Storge' , Product_name = Product_name)
            if  not valid : 
                self.cursor1.execute(f"""UPDATE Storge SET [Product_name] = '{Product_name}', [buying_price] = '{buying_price}', [sell_price] = {sell_price}, 
                                     [unit_type] = '{unit_type}' , [total_num_of_units] = '{total_num_of_units}' where Product_name like '%{Product_name}%' ;""")
                self.connection.commit() 
                QMessageBox.information(self, "i" , "تم تجديد البيانات بنجاح ")
                self.Storge_ShowAll_func()
            else: QMessageBox.warning(self, "i" , "لا يوجد منتج بهذا الاسم") 
        else: QMessageBox.warning(self, "i" , " لم تقم بأدخال أحدي البيانات  ")
    
    def Storge_ShowAll_func(self):
        self.cursor1.execute(f"""select * from Storge""")
        rows = self.cursor1.fetchall()
        self.Storge_Table.clearContents()
        if rows:
                    self.Storge_Table.setRowCount(len(rows))
                    for i, row in enumerate(rows):
                        for j in range(len(row)):
                            item = QTableWidgetItem(str(row[j]))
                            self.Storge_Table.setItem(i, j, item)
    
    def get_data_from_storge_Ui(self):
        Product_name = self.storge_tab_product_name_line.text()
        buying_price = self.storge_tab_buy_price_line.text()
        sell_price = self.storge_tab_unit_price_line.text()
        total_num_of_units = self.storge_tab_total_numof_unit_line.text()
        unit_type  = self.storge_tab_unit_type_comboBox.currentText()
        return Product_name , buying_price , sell_price , unit_type, total_num_of_units 
 
  
#*------------------------------------------------------------------------------ selling tab  --------------------------------------------
        
    def Handel_sell_Buttons(self):
        self.selling_tab_sell_btn.clicked.connect(self.selling_func)
        self.selling_tab_search_btn.clicked.connect(self.selling_Search_func)
        self.selling_tab_showall_btn.clicked.connect(self.selling_ShowAll_func)
        self.storge_tab_add__to_selltable_btn.clicked.connect(self.add_to_small_table)
        self.delete_data_from_selling_small_table()
        self.storge_tab_delete_selltable_btn.clicked.connect(self.delete_from_small_table_func)
        self.storge_tab_print_selltable_btn.clicked.connect(self.print_from_smallTable_func)
        
    def delete_data_from_selling_small_table(self) :
        self.cursor1.execute(""" DELETE from small_sell_table; """)
        self.connection.commit()
    
    def selling_func(self):
        self.cursor1.execute(f"""select * from small_sell_table""")
        rows = self.cursor1.fetchall()  
        total_amount_of_all_invoice = 0       
        list_of_products_names = ''
        today_date = dt.today().date()
        if rows:
            for row in rows :
                self.cursor1.execute(f"""select total_num_of_units from Storge where Product_name like '%{row[0]}%' """)
                num_of_units = self.cursor1.fetchall()
                
                num_of_units = int(num_of_units[0][0])
                if num_of_units > 0 and num_of_units >= row[2] :
                    self.cursor1.execute(f""" UPDATE Storge SET [total_num_of_units] = '{num_of_units - int(row[2]) }' where [Product_name] = '{row[0]}';""")
                    self.connection.commit()
                    total_amount_of_all_invoice += row[3]
                    list_of_products_names+=("  ||  "+str(row[0]+f'بكميه = {row[2]}'))
                else : 
                    QMessageBox.information(self, "i" , "العدد المتاح من المنتج في المخزن غير كافي لعمليه البيع ") 
                    return 
            self.add_invoice_to_customer( today_date  , total_amount_of_all_invoice  , list_of_products_names)
            self.small_sell_tble.clearContents()
            self.delete_from_small_table_func()
            self.sellingTab_total_monay_of_invoice.clear()
            self.selling_ShowAll_func()
                                          
    def add_invoice_to_customer(self , today_date  , total_amount_of_all_invoice  , list_of_products_names):
        customer_name = self.selling_tab_customer_name_line.text()
        paid_amount =self.sellingTab_paid_amount_invoice_2.text()
        self.add_to_invoice_history(today_date  , total_amount_of_all_invoice  , list_of_products_names , customer_name)
        
        self.cursor1.execute(f"""select * from customers where name = '{customer_name}' """)
        customer_info = self.cursor1.fetchone()
        if customer_info:
            print(customer_info[2])
            if int(customer_info[2]) == int(0):
                print('111111111111111111')
                self.reset_customer_info(customer_name)  
                remaend_amount = float(total_amount_of_all_invoice) - float(paid_amount)
                self.cursor1.execute(f""" UPDATE customers SET [total_amount] = {total_amount_of_all_invoice} , [remaend_amount] = {remaend_amount} where [name] = '{customer_name}';""")
                self.connection.commit()
                
            else:
                print('22222222222222222222')
                curr_invoice_remaend_amount = float(total_amount_of_all_invoice) - float(paid_amount)
                new_remaend_amount = curr_invoice_remaend_amount + customer_info[2]
                new_total_amount = total_amount_of_all_invoice + customer_info[1]
                self.cursor1.execute(f""" UPDATE customers SET [total_amount] = {new_total_amount} , [remaend_amount] = {new_remaend_amount} where [name] = '{customer_name}';""")
                self.connection.commit()
                
        else:
            print('3333333333333333333')
            remaend_amount = float(total_amount_of_all_invoice) - float(paid_amount)
            self.cursor1.execute(""" insert into customers ( [name] , [total_amount], [remaend_amount] ) 
                                        values (?,?,?)""", ( today_date  , total_amount_of_all_invoice  , remaend_amount ))
            self.connection.commit() 
            QMessageBox.information(self, "i" , "تم  العمليه بنجاح")
            
    def add_to_invoice_history(self ,today_date  , total_amount_of_all_invoice  , list_of_products_names ,customer_name  )   :
        self.cursor1.execute(""" insert into invoice_history ( [date] , [total_amount], [list_of_products_names] ) 
                                        values (?,?,?)""", ( today_date  , total_amount_of_all_invoice  , list_of_products_names  ))
        self.connection.commit() 
                    
    def reset_customer_info(self  , customerName):
        self.cursor1.execute(f""" UPDATE customers SET [total_amount] = {0} where [name] = '{customerName}';""")
        self.connection.commit()
    
    def selling_Search_func(self):
        name = self.selling_tab_product_name_line.text()
        if name:
            self.cursor1.execute(f"""select * from Storge where Product_name like '%{name}%'""")
            rows = self.cursor1.fetchall()
            self.selling_tableWidget.clearContents()
            if rows:
                self.selling_tableWidget.setRowCount(len(rows))
                for i, row in enumerate(rows):
                    for j in range(len(row)):
                        item = QTableWidgetItem(str(row[j]))
                        self.selling_tableWidget.setItem(i, j, item)
            else:
                QMessageBox.warning(self, "i" , "لا يوجد منتج بهذه المعلومات")
        else:
            QMessageBox.warning(self, "i" , "يجب ادخال اسم المنتج لأتمام عمليه البحث")

    def selling_ShowAll_func(self):
        self.cursor1.execute(f"""select * from Storge""")
        rows = self.cursor1.fetchall()
        self.selling_tableWidget.clearContents()
        if rows:
                    self.selling_tableWidget.setRowCount(len(rows))
                    for i, row in enumerate(rows):
                        for j in range(len(row)):
                            item = QTableWidgetItem(str(row[j]))
                            self.selling_tableWidget.setItem(i, j, item)
    
    def selling_get_data_from_Ui(self):
        Product_name = self.storge_tab_product_name_line.text()
        buying_price = self.storge_tab_buy_price_line.text()
        sell_price = self.storge_tab_unit_price_line.text()
        total_num_of_units = self.storge_tab_total_numof_unit_line.text()
        unit_type  = self.storge_tab_unit_type_comboBox.currentText()
        return Product_name , buying_price , sell_price , unit_type, total_num_of_units 
  
    def show_all_data_in_small_table(self):
        self.cursor1.execute(f"""select * from small_sell_table""")
        rows = self.cursor1.fetchall()
        self.small_sell_tble.clearContents()
        if rows:
                    self.small_sell_tble.setRowCount(len(rows))
                    for i, row in enumerate(rows):
                        for j in range(len(row)):
                            item = QTableWidgetItem(str(row[j]))
                            self.small_sell_tble.setItem(i, j, item)
    
    def add_to_small_table(self):
        Product_name  , sell_price , total_num_of_units = self.sellingTab_get_data_from_Ui() 
        if Product_name  and sell_price and total_num_of_units:
            if not self.check_if_productname_is_Available(TableName='Storge' , Product_name=Product_name):
                total_monay = int(sell_price)  * int(total_num_of_units)
                Mainapp.total_monay_of_invoice += total_monay
                self.cursor1.execute(""" insert into small_sell_table ( [product_name] , [price] , [number_of_units] , [total_monay]  ) 
                                            values (?,?,?,?)""", ( Product_name  , sell_price  , total_num_of_units  , total_monay))
                self.connection.commit() 
                total_invoice_amount = self.calc_total_invoice_amount()
                self.sellingTab_total_monay_of_invoice.setText(str(total_invoice_amount[0]))
                self.show_all_data_in_small_table()
            else : QMessageBox.warning(self,'i'," لا يوجد منتج بهذا الاسم في المخزن  ")
        else :QMessageBox.warning(self,'i',"تاكد من ادخال كل البيانات ")
        
    def calc_total_invoice_amount(self):
        self.cursor1.execute(""" select SUM(total_monay) from small_sell_table """)
        total_invoice_amount = self.cursor1.fetchone() 
        return total_invoice_amount 
    
    def sellingTab_get_data_from_Ui(self):
        product_name  = self. selling_tab_product_name_line.text()
        selling_price = self.storge_tab_buy_price_line_2.text()
        total_num_of_units = self.storge_tab_total_numof_unit_line_2.text()
        return product_name  , selling_price  , total_num_of_units
  
    def Handel_seelingTab_tabels_Ui(self):
        self.selling_tableWidget.setColumnWidth(0,270)
        self.selling_tableWidget.setColumnWidth(1,270)
        self.selling_tableWidget.setColumnWidth(2,270)
        self.selling_tableWidget.setColumnWidth(3,270)
        self.selling_tableWidget.setColumnWidth(4,270)
        
        self.small_sell_tble.setColumnWidth(0,140)
        self.small_sell_tble.setColumnWidth(1,140)
        self.small_sell_tble.setColumnWidth(2,140)
        self.small_sell_tble.setColumnWidth(3,140)
        
        self.invoice_table.setColumnWidth(0,450)
        self.invoice_table.setColumnWidth(1,450)
        self.invoice_table.setColumnWidth(2,450)
        
    def delete_from_small_table_func(self):
        name = self.selling_tab_product_name_line.text()
        self.cursor1.execute(f"""delete from small_sell_table WHERE product_name like '%{name}%' """)
        self.connection.commit()
        self.show_all_data_in_small_table()
        total_invoice_amount = self.calc_total_invoice_amount()
        self.sellingTab_total_monay_of_invoice.setText(str(total_invoice_amount[0]))
        
    def print_from_smallTable_func(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Retrieve data from the QTableWidget
        for row in range(self.small_sell_tble.rowCount()):
            for column in range(self.small_sell_tble.columnCount()):
                item = self.small_sell_tble.item(row, column)
                if item is not None:
                    value = item.text()
                    sheet.cell(row=row+1, column=column+1, value=value)

        # Save the workbook to a file
        n = random.randint(1,1000000)
        today_date = dt.today().date()
        workbook.save(f'invoice({n}) {today_date}.xlsx')

#*------------------------------------------------------------------------------ customer tab  --------------------------------------------
    def Handel_customers_tab_btns(self):
        self.customer_tab_add_btn.clicked.connect(self.customer_tab_add_func)
        self.customer_tab_search_btn.clicked.connect(self.customer_tab_search_func)
        self.customer_tab_showall_btn.clicked.connect(self.customer_tab_showAll_func)
        self.customer_tab_delete_btn.clicked.connect(self.customer_tab_delete_func)
        self.customer_tab_UPDATE_btn.clicked.connect(self.customer_tab_update_func)

    def customers_tables_ui(self):
        self.tableWidget_28.setColumnWidth(0,450)
        self.tableWidget_28.setColumnWidth(1,450)
        self.tableWidget_28.setColumnWidth(2,450)
    
    def customer_tab_add_func(self):
        customer_name , total_amount , remaining_amount  = self.customers_get_data() 
        if customer_name and remaining_amount >= 0:
            if  self.check_if_customerName_available(customer_name):
                self.cursor1.execute(""" insert into customers ( [name] , [total_amount] , [remaend_amount]  ) 
                                            values (?,?,?)""", ( customer_name  , total_amount  , remaining_amount))
                self.connection.commit() 
                self.customer_tab_showAll_func()
            else : QMessageBox.warning(self,'i',"يوجد عميل بهذا الاسم ادخل اسم مختلف ")
        else :QMessageBox.warning(self,'i',"تاكد من ادخال كل البيانات ")
    
    def customer_tab_search_func(self):
        name = self.customers_tab_Name_LINE.text()
        if name:
            self.cursor1.execute(f"""select * from customers where name like '%{name}%'""")
            rows = self.cursor1.fetchall()
            self.tableWidget_28.clearContents()
            if rows:
                self.tableWidget_28.setRowCount(len(rows))
                for i, row in enumerate(rows):
                    for j in range(len(row)):
                        item = QTableWidgetItem(str(row[j]))
                        self.tableWidget_28.setItem(i, j, item)
            else:
                QMessageBox.warning(self, "i" , "لا يوجد عميل بهذه المعلومات")
        else:
            QMessageBox.warning(self, "i" , "يجب ادخال اسم العميل لأتمام عمليه البحث")
    
    def customer_tab_showAll_func(self):
            self.cursor1.execute(f"""select * from customers  """)
            rows = self.cursor1.fetchall()
            self.tableWidget_28.clearContents()
            if rows:
                self.tableWidget_28.setRowCount(len(rows))
                for i, row in enumerate(rows):
                    for j in range(len(row)):
                        item = QTableWidgetItem(str(row[j]))
                        self.tableWidget_28.setItem(i, j, item)
    
    def customer_tab_delete_func(self):
        customer_name = self.customers_tab_Name_LINE.text()
        self.cursor1.execute(f"""delete from customers WHERE name like '%{customer_name}%' """)
        self.connection.commit()
        self.customer_tab_showAll_func()
    
    def customers_get_data(self):
        customer_name = self.customers_tab_Name_LINE.text()
        total_amount  = self.customers_tab_total_amount_LINE.text()
        paid_amount   = self.customers_tab_paid_amount_LINE.text()
        
        try:
            remaining_amount = float(total_amount)  - float(paid_amount)
            return customer_name ,total_amount , remaining_amount
        except Exception :
            QMessageBox.warning(self,'i','حدث خطأ في العمليه ')
        
    def check_if_customerName_available(self , name):
        self.cursor1.execute(f"""select * from customers where name = '{name}' """)
        rows = self.cursor1.fetchall()
        if rows:
            return False
        else:
            return True

    def customer_tab_update_func(self):
        try:
            customer_name = self.customers_tab_Name_LINE.text()
            paid_amount = self.customers_tab_paid_amount_LINE.text()
            remening_amount = self.get_remining_amout(customer_name)
            new_remening_amount = remening_amount - float(paid_amount)
            if new_remening_amount >= 0 :
                self.cursor1.execute(f""" UPDATE customers SET [remaend_amount] = {new_remening_amount} where [name] = '{customer_name}';""")
                self.connection.commit()
                QMessageBox.information(self , 'i' , 'تمت العمليه بنجاح ')
            else:
                QMessageBox.warning(self,'i','المبلغ المدفوع اكبر من المبلغ المتبقي علي العميل ')
        except Exception as e:
             print(e)
             QMessageBox.warning(self,'i','حدث خطأ في العمليه ')
        
    def get_remining_amout(self , customer_name):
        self.cursor1.execute(f"""select remaend_amount from customers where name like '%{customer_name}%' """)
        remaend_amount = self.cursor1.fetchone()
        return float(remaend_amount[0])
        
    

#*------------------------------------------------------------------------------ invoice tab  --------------------------------------------
    def Handel_invoice_tab_btns(self):
        self.storge_tab_add_btn_5.clicked.connect(self.invoice_tab_search_func)
        self.storge_tab_showall_btn_4.clicked.connect(self.invoice_tab_Showall)
        self.storge_tab_print_btn_5.clicked.connect(self.invoice_tab_print_func)
        
        
    def invoice_tab_Showall(self):
        self.cursor1.execute(f"""select * from invoice_history""")
        rows = self.cursor1.fetchall()
        self.invoice_table.clearContents()
        if rows:
                    self.invoice_table.setRowCount(len(rows))
                    for i, row in enumerate(rows):
                        for j in range(len(row)):
                            item = QTableWidgetItem(str(row[j]))
                            self.invoice_table.setItem(i, j, item)
                            
    def invoice_tab_search_func(self):
        invoive_date = self.invoice_tab_data_lineedit.text()
        self.cursor1.execute(f"""select * from invoice_history where date like '%{invoive_date}%'""")
        rows = self.cursor1.fetchall()
        self.invoice_table.clearContents()
        if rows:
                    self.invoice_table.setRowCount(len(rows))
                    for i, row in enumerate(rows):
                        for j in range(len(row)):
                            item = QTableWidgetItem(str(row[j]))
                            self.invoice_table.setItem(i, j, item)
    
    def invoice_tab_print_func(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Retrieve data from the QTableWidget
        for row in range(self.invoice_table.rowCount()):
            for column in range(self.invoice_table.columnCount()):
                item = self.invoice_table.item(row, column)
                if item is not None:
                    value = item.text()
                    sheet.cell(row=row+1, column=column+1, value=value)

        # Save the workbook to a file
        n = random.randint(1,1000000)
        today_date = dt.today().date()
        workbook.save(f'invoice({n}) {today_date}.xlsx')

#* ------------------------------------------------------------------------ support functions ------------------------------------------------------------------------------
    def check_if_productname_is_Available(self,TableName , Product_name):
            print(TableName)
            print(Product_name)
            self.cursor1.execute(f"""select * from {TableName} where Product_name like '%{Product_name}%' ;""")
            code = self.cursor1.fetchone()
            if code :
                return False
            else:
                return True 
        
     
   
   
#* ------------------------------------------------------------------------ get data from selected row ------------------------------------------------------------------------------
    def get_data_from_selected_row (self):
        self.Storge_Table.currentItemChanged.connect(self.on_item_changed_Storge_table)
        self.selling_tableWidget.currentItemChanged.connect(self.on_item_changed_selling_tableWidget)
        self.tableWidget_28.currentItemChanged.connect(self.on_item_changed_customer_table)
        self.small_sell_tble.currentItemChanged.connect(self.on_item_changed_SellSmall_table)
        
    def on_item_changed_selling_tableWidget(self, current):
        if current is not None:
            row = current.row()
            data = []
            for col in range(self.selling_tableWidget.columnCount()):
                item = self.selling_tableWidget.item(row, col)
                data.append(item.data(Qt.DisplayRole))
            #print(data)
            self.selling_tab_product_name_line.setText(data[0])
            self.storge_tab_buy_price_line_2.setText(data[2])
                       
    def on_item_changed_Storge_table(self, current):
        if current is not None:
            row = current.row()
            data = []
            for col in range(self.Storge_Table.columnCount()):
                item = self.Storge_Table.item(row, col)
                data.append(item.data(Qt.DisplayRole))
            #print(data)
            self.storge_tab_product_name_line.setText(data[0])
            self.storge_tab_buy_price_line.setText(data[1])
            self.storge_tab_unit_price_line.setText(data[2])
            self.storge_tab_total_numof_unit_line.setText(data[4])
            self.storge_tab_unit_type_comboBox.setCurrentText(data[3])
   
    def on_item_changed_customer_table(self, current):
        if current is not None:
            row = current.row()
            data = []
            for col in range(self.tableWidget_28.columnCount()):
                item = self.tableWidget_28.item(row, col)
                data.append(item.data(Qt.DisplayRole))
            #print(data)
            self.customers_tab_Name_LINE.setText(data[0])
            self.customers_tab_total_amount_LINE.setText(data[1])
     
    def on_item_changed_SellSmall_table(self, current):
        if current is not None:
            row = current.row()
            data = []
            for col in range(self.small_sell_tble.columnCount()):
                item = self.small_sell_tble.item(row, col)
                data.append(item.data(Qt.DisplayRole))
            #print(data)
            self.selling_tab_product_name_line.setText(data[0])
            self.storge_tab_buy_price_line_2.setText(data[1])
            
           
    
    def Handel_main_buttons(self):
        self.open_storge_tab__btn.clicked.connect(self.open_Storge_tab)   
        self.open_sell_tab__btn.clicked.connect(self.open_Sell_tab) 
        self.open_invoice_tab__btn.clicked.connect(self.open_selling_History_tab) 
        self.open_customers_tab__btn.clicked.connect(self.open_customers_tab) 
        
    def open_Storge_tab(self):
        self.tabWidget.setCurrentIndex(0)        
    def open_Sell_tab(self):
        self.tabWidget.setCurrentIndex(1)  
    def open_selling_History_tab(self):
        self.tabWidget.setCurrentIndex(2)       
    def open_customers_tab(self):
        self.tabWidget.setCurrentIndex(3)
        
      

def main():
    app = QApplication(sys.argv)
    window = Mainapp()
    window.show()
    app.exec_()


if __name__ == "__main__":
    main()